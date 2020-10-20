#!/usr/bin/python3

import os

from openpyxl import load_workbook

from models.base import db_session
from models.category import Category
from settings import ROOT_PATH


def fetch_categories(file, sheet):
    """ 从 database.xlsx 读取数据 """
    wb = load_workbook(filename=file)
    ws = wb[sheet]
    for row in ws['A2:C{}'.format(ws.max_row)]:
        values = (name, slug, memo) = [cell.value for cell in row]
        if name:
            category = Category(**dict(zip(("name", "slug", "memo"), values)))
            db_session.add(category)
    db_session.commit()
    print("插入了 %s 行数据到 categories 表." % str(ws.max_row - 1))


if __name__ == "__main__":
    Category.__table__.drop(checkfirst=True)
    Category.__table__.create()
    fetch_categories(os.path.join(ROOT_PATH, "basedata/database.xlsx"),
                     "categories")
