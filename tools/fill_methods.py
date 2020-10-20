#!/usr/bin/python3

import os

from openpyxl import load_workbook

from models.base import db_session
from models.qc_method import QCMethod
from settings import ROOT_PATH


def fetch_methods(file, sheet):
    """ 从 database.xlsx 读取数据 """
    wb = load_workbook(filename=file)
    ws = wb[sheet]
    for row in ws['A2:C{}'.format(ws.max_row)]:
        values = (name, file, content) = [cell.value for cell in row]
        if name:
            method = QCMethod(**dict(zip(("name", "file", "content"), values)))
            db_session.add(method)
    db_session.commit()
    print("插入了 %s 行数据到 methods 表." % str(ws.max_row - 1))


if __name__ == "__main__":
    QCMethod.__table__.drop(checkfirst=True)
    QCMethod.__table__.create()
    fetch_methods(os.path.join(ROOT_PATH, "basedata/database.xlsx"),
                  "methods")
