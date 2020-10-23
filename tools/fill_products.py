#!/usr/bin/python3

import os

from openpyxl import load_workbook

from models.base import db_session
from models.category import Category
from models.product import Product
from report.common import get_products_form_xls
from settings import ROOT_PATH


def fill_product_sheet(filepath):
    """填充产品类别"""
    wb = load_workbook(filepath)
    ws = wb['products']
    for row_index in range(2, ws.max_row + 1):
        market_name = ws['E{}'.format(row_index)].value
        '''
        color = ws['J{}'.format(row_index)].value

        if market_name.find("BL") >= 0:
            if not color:
                ws['J{}'.format(row_index)] = "蓝色"

        if market_name.find("BK") >= 0:
            if not color:
                ws['J{}'.format(row_index)] = "黑色"

        if market_name.find("WL") >= 0 or market_name.find("WB") >= 0:
            if not color:
                ws['J{}'.format(row_index)] = "白色"

        if market_name.find("G") >= 0 or market_name.find("SP") >= 0:
            if not color:
                ws['J{}'.format(row_index)] = "绿色"
        '''

        kind = ws['B{}'.format(row_index)].value
        category = ws['F{}'.format(row_index)].value

        if kind and not category:
            if kind.startswith('a2000'):
                ws['F{}'.format(row_index)] = "A-2000"
            elif kind.startswith('a2100'):
                ws['F{}'.format(row_index)] = "A-2100"
            elif kind.startswith('a9000'):
                ws['F{}'.format(row_index)] = "A-9000"
            elif kind.startswith('a9060a'):
                ws['F{}'.format(row_index)] = "A-9060A"
            elif kind.startswith('k2500'):
                ws['F{}'.format(row_index)] = "K-2500"
            elif market_name.startswith('SP'):
                ws['F{}'.format(row_index)] = "SPXX"
            elif kind.startswith('h9100b'):
                ws['F{}'.format(row_index)] = "H-8100B/H-9100B"
            elif kind.startswith('h8100'):
                ws['F{}'.format(row_index)] = "H-8100"
            elif kind.startswith('h9100'):
                ws['F{}'.format(row_index)] = "H-9100"
            elif kind.startswith('ts3000'):
                ws['F{}'.format(row_index)] = "TS-3000"
            elif kind.startswith('tm3100'):
                ws['F{}'.format(row_index)] = "TM-3100"
            elif kind.startswith('uvm1800'):
                ws['F{}'.format(row_index)] = "UVM-1800"
            elif kind.startswith('uvs1000'):
                ws['F{}'.format(row_index)] = "UVS-1000"
            elif kind.startswith('p2700'):
                ws['F{}'.format(row_index)] = "P-2700"
            else:
                ws['F{}'.format(row_index)] = "undefined"

    wb.save(filepath)


def fetch_products(filepath):
    category, products = get_products_form_xls(filepath)
    for product in products:
        if product["product"]:
            product = Product(**product)
            category = Category.query.filter_by(slug=category).first()
            product.category = category
            db_session.add(product)

    db_session.commit()
    print("插入了 %d 行数据到 products 表." % len(products))


if __name__ == "__main__":
    database_filepath = os.path.join(ROOT_PATH, "basedata/database.xlsx")
    fill_product_sheet(database_filepath)

    Product.__table__.drop(checkfirst=True)
    Product.__table__.create()
    fetch_products(database_filepath)
