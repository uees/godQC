import os
import random
from datetime import datetime

from openpyxl import load_workbook
from settings import REPORTS_ROOT


def today_reports_root() -> str:
    """ 返回当日报告文件夹路径 """
    today_dir_name = datetime.strftime(datetime.now(), '%Y-%m-%d')
    today_path = os.path.join(REPORTS_ROOT, today_dir_name)
    if not os.path.exists(today_path):
        os.mkdir(today_path)
    return today_path


def unique_filepath(path: str, filename: str) -> str:
    """ 检查并生成唯一的文件名 """
    filepath = os.path.join(path, filename)

    if os.path.exists(filepath):
        shotname, extension = os.path.splitext(filename)
        filename = '%s_.%s' % (shotname, extension)

        return unique_filepath(path, filename)

    return filepath


def get_products_form_xls(filepath: str) -> (str, list):
    products = []
    wb = load_workbook(filepath)
    ws = wb["products"]
    for row in ws['A2:J{}'.format(ws.max_row)]:
        internal_name, kind, label_viscosity, viscosity_width, market_name, category, part_a, part_b, ab_ratio, color = [cell.value for cell in row]
        values = (internal_name, market_name, part_a, part_b, ab_ratio, color, label_viscosity, viscosity_width)
        product = dict(zip(("internal_name", "market_name", "part_a", "part_b", "ab_ratio", "color", "label_viscosity", "viscosity_width"), values))
        products.append(product)

    return category, products


def write_products_xls(filepath, sheetname, products):
    wb = load_workbook(filepath)
    ws = wb[sheetname]
    row = 2
    for product in products:
        ws[f'A{row}'] = product.internal_name
        ws[f'B{row}'] = product.category.slug
        ws[f'C{row}'] = product.label_viscosity
        ws[f'D{row}'] = product.viscosity_width
        ws[f'E{row}'] = product.market_name
        ws[f'F{row}'] = product.category.slug
        ws[f'G{row}'] = product.part_a
        ws[f'H{row}'] = product.part_b
        ws[f'I{row}'] = product.ab_ratio
        ws[f'J{row}'] = product.color

        row += 1

    wb.save(filename=filepath)


def make_viscosity(viscosity_limit):
    """根据粘度范围生成一个假粘度值"""
    if '±' in viscosity_limit:
        mid, limit = viscosity_limit.split('±')
        mid, limit = int(mid), int(limit)

        if limit > 15:
            limit = 15

        return f'{random.randint(mid-limit, mid+limit)}'

    if '~' in viscosity_limit:
        viscosity_limit = viscosity_limit.replace('~', '-')

    if '-' in viscosity_limit:
        min_, max_ = viscosity_limit.split('-')
        min_, max_ = int(min_), int(max_)

        return f'{random.randint(min_, max_)}'


def make_viscosity_limit(product):
    mid, limit = product.label_viscosity, product.viscosity_width  # str
    if mid and limit:
        mid, limit = int(mid), int(limit)
        viscosity_limit = f"{mid}±{limit}"
        return viscosity_limit, make_viscosity(viscosity_limit)

    return '', ''


def null2str(value):
    if value is None:
        value = ''
    return value
